"""
Super Game Seeder — seeds all providers, categories, and games.

Categories are inferred from each game (e.g. Roulette, Blackjack, Baccarat,
Slots, Sports Betting …) and created dynamically — no fixed set of categories.

Providers covered:
  • Ezugi       — embedded list + docs/games/ezugiwebp images
  • JILI        — embedded list + docs/games/jiliwebp images
  • Evolution Live  — docs/games/Evolution live.xlsx + docs/games/evolutionwebp
  • Evoplay Asia    — docs/games/evoplay asia.xlsx   (no images folder)
  • Pragmatic Live  — docs/games/Pragmatic live.xlsx + docs/games/pragmaticlivewebp
  • SABA Sports     — docs/games/SABA Sports.xlsx    (no images folder)
  • Sexy Gaming     — docs/games/Sexy Gaming.xlsx    + docs/games/sexygamingwebp
  • SmartSoft Gaming — docs/games/SmartSoft Gaming.xlsx (no images folder)
  • Spribe          — docs/games/spribe.txt           + docs/games/spribe
  • LuckSportsGaming — docs/games/lucksportsgaming.txt (no images folder)

Options:
  --dry-run      : print stats only; no DB writes
  --fresh        : delete all seeded providers/games first, then re-seed
  --full-reset   : delete ALL Game, GameCategory, GameProvider, then re-seed
  --providers    : comma-separated provider codes to limit (e.g. ezugi,jili,spribe)
  --images-only  : only fill missing images; do not create new games

Path: Game data and images are read from a single folder. Resolved in order:
  (1) DOCS_GAMES_PATH env or settings, (2) commands/games_data (shipped with app for VPS),
  (3) BASE_DIR.parent/docs/games, (4) BASE_DIR/docs/games. Put XLSX, TXT, and image folders in games_data so the seeder works on VPS without config.
"""

from __future__ import annotations

import os
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand

from core.models import Game, GameCategory, GameProvider
from core.management.utils import (
    find_image_for_game_in_folders,
    get_image_folder_candidates,
)


def _resolve_docs_games_path() -> Path:
    """
    Resolve game data directory. Order: (1) DOCS_GAMES_PATH if set,
    (2) commands/games_data if it exists (shipped with app for VPS),
    (3) BASE_DIR.parent/docs/games, (4) BASE_DIR/docs/games.
    """
    base = Path(settings.BASE_DIR)
    explicit = os.environ.get("DOCS_GAMES_PATH") or getattr(settings, "DOCS_GAMES_PATH", None)
    if explicit:
        return Path(explicit).resolve()
    # Same package as this command — deploy with app so VPS works without config
    commands_dir = Path(__file__).resolve().parent
    games_data = commands_dir / "games_data"
    if games_data.exists():
        return games_data
    parent_docs = base.parent / "docs" / "games"
    inner_docs = base / "docs" / "games"
    if parent_docs.exists():
        return parent_docs
    if inner_docs.exists():
        return inner_docs
    return games_data  # default: use games_data so deploy can add files there


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

DOCS_GAMES = _resolve_docs_games_path()

# ---------------------------------------------------------------------------
# Category inference — each game's category name is derived from its name/provider
# ---------------------------------------------------------------------------

def infer_subcategory(name: str, provider_code: str) -> str:
    """Return a category name inferred from game name and provider."""
    n = (name or "").lower()
    p = (provider_code or "").lower()

    # Sports
    if p in ("saba_sports", "lucksportsgaming", "lucksportgaming"):
        return "Sports Betting"

    # Fishing
    if "fishing" in n:
        return "Fishing"

    # Bingo
    if "bingo" in n:
        return "Bingo"

    # Roulette
    if "roulette" in n:
        return "Roulette"

    # Keno
    if "keno" in n:
        return "Keno"

    # Crash / instant
    if any(kw in n for kw in ("crash", "aviator", "go rush", "balloon")):
        return "Crash"

    # Mines / instant games
    if any(kw in n for kw in ("mines", "plinko", "limbo", "tower", "wheel")):
        return "Instant Games"

    # Blackjack
    if "blackjack" in n:
        return "Blackjack"

    # Baccarat
    if "baccarat" in n or "bac bo" in n:
        return "Baccarat"

    # Poker / card games
    if any(kw in n for kw in ("poker", "hold'em", "holdem", "stud", "caribbean", "mini flush")):
        return "Poker"

    # Teen Patti / Andar Bahar / South-Asian table games
    if any(kw in n for kw in ("teen patti", "teenpatti", "andar bahar", "jhandi munda", "32 cards", "7up7down", "ak47", "hilo", "hi lo")):
        return "Table Games"

    # Sic Bo / Dice
    if any(kw in n for kw in ("sic bo", "sicbo", "craps", "dice")):
        return "Dice Games"

    # Dragon Tiger
    if "dragon tiger" in n or "dragon & tiger" in n:
        return "Dragon Tiger"

    # Game-show style live
    if any(kw in n for kw in ("dream catcher", "monopoly", "crazy time", "funky time", "mega ball", "deal or no deal", "gonzo", "football studio", "stock market", "cash or crash", "crazy coin", "crazy pachinko", "imperial quest", "bac bo")):
        return "Game Shows"

    # Lucky / scratch / other live
    if any(kw in n for kw in ("lucky 7", "cricket war", "bet on numbers", "color game", "color prediction", "sedie", "fish prawn crab", "thai hi lo", "thai fish")):
        return "Table Games"

    if any(kw in n for kw in ("speed", "lightning", "immersive")):
        return "Live Casino"

    # Rummy / Ludo / Call break / Tongits — card/board games
    if any(kw in n for kw in ("rummy", "ludo", "callbreak", "tongits", "pusoy", "pool rummy")):
        return "Card & Board Games"

    # Lobby entries
    if "lobby" in n:
        return "Lobby"

    return "Slots"


# ---------------------------------------------------------------------------
# Embedded game lists
# ---------------------------------------------------------------------------

EZUGI_GAMES: list[tuple[str, str]] = [
    ("Blackjack C", "cd0e742af59d62f2241c1f6bc19954c5"),
    ("Blackjack B", "90bd51cc8cde4d06bbb6ae787c8c3eb3"),
    ("EZ Dealer Roleta Brazileira", "4165dec80667c631a66941c68a5bee96"),
    ("Roleta da sorte", "eacbc601b30b2992db7c3eda2a777fe6"),
    ("Dragon Tiger da Sorte", "f84c5ce9ae53fac2e7afa9f8157e453c"),
    ("Blackjack A", "d9fc983ad1ac44e2a6365ae6cc5c9762"),
    ("Blackjack da Sorte", "4f22281594a261d99c1b1222bc2d3a8a"),
    ("Oracle 360 Roulette", "c74c90c712566b3212cd08a4c191275d"),
    ("Oracle Real Roulette", "c5d4fd6cec78dd439ed2ee33c8965777"),
    ("Russian Roulette", "ce1b314dccf3756a581d117190ddd172"),
    ("Turkish Roulette", "f8bbea8c1a3b2204190d6a7e3c8d55e8"),
    ("EZ Dealer Roulette Mandarin", "db62931938bcaf0b327b11304a406b16"),
    ("EZ Dealer Roulette Japanese", "4a963476f45508711a7147ba888600ad"),
    ("EZ Roulette", "fb53959d1f55434d555ee50e0fc764b8"),
    ("Fiesta Roulette", "4e94d574914b472cb4ecc4f3c05647d4"),
    ("Ruleta del Sol", "4589bec2f464797bb0752d2eb283babd"),
    ("Spanish Roulette", "7d0d91d4477b9d14e3a4ba40e34451ea"),
    ("Casino Marina Roulette 2", "44b99989a409c0ca24aca784f0433dcc"),
    ("Casino Marina Roulette 1", "d205c518208016404504e995620d2b83"),
    ("Ultimate Andar Bahar", "75f81c56555d394503f544f3431ef370"),
    ("Teen Patti 3 Card", "26f9f76a8fc813b8abcb6b8cb03c2eab"),
    ("Namaste Roulette", "b1ffb1afd5b76785bd4ee21e31400849"),
    ("Prestige Auto Roulette", "efaed662fbebbb84e056c09580ae1aa4"),
    ("Diamond Roulette", "a40c7e3222a17717bcc1d2e4f5d6eae8"),
    ("Speed Auto Roulette", "f4299915859041e94b641a558a1ca9df"),
    ("Speed Roulette", "b5c8e49fdd80b57de6da0e234b1bd683"),
    ("Fiesta Baccarat", "66a525d29fdfb01af4335ceabfb0cad2"),
    ("Casino Marina Baccarat 4", "527db204952a306f8459b9d702dfb285"),
    ("Casino Marina Baccarat 3", "243c511540c8d82597245bd282c327a1"),
    ("Casino Marina Baccarat 2", "58501dadbf6088c4722e72660a1f38b7"),
    ("Turkish Unlimited Blackjack", "5b66adfdde56956cf3d4273acfad99d4"),
    ("Fiesta Unlimited Blackjack", "5255de9f809be3a1515cf28095a95039"),
    ("Spanish Unlimited Blackjack", "e34d828be9c5dbd861dbcc414d2daad7"),
    ("Turkish Blackjack 3", "7b4f503301dd3fbbb0beb71b814452c9"),
    ("Russian Blackjack 2", "716e3dd6a3e560067425fec4951abe25"),
    ("Russian Blackjack 1", "4244969212746e76beb92f71ba300114"),
    ("Turkish Blackjack 2", "25384176a2560a0a69c301b1c8cf83f9"),
    ("Turkish Blackjack 1", "9e8ce809e74cc1ebda5ca59a927def6b"),
    ("Rumba Blackjack 4", "1c77a06449c384b97f6239572ef87be3"),
    ("Rumba Blackjack 3", "e6a2e3cc081f28298164d9197c38ec7c"),
    ("Rumba Blackjack 2", "22c57f788355265137a61874d0b53bb9"),
    ("Rumba Blackjack 1", "8ef17e9b4c5c67b7f43f4bced3c31a27"),
    ("VIP Diamond Blackjack", "48420430ede7a5d7615dae19aa4463a7"),
    ("Surrender Blackjack", "d9723621d4007265d66cc115b5a953df"),
    ("Gold Blackjack 6", "9353bede98efba162ed5b04534e9ac00"),
    ("Blackjack 7", "3ab6fd647eac8e687af18ce5bceadfa5"),
    ("Romanian Blackjack", "7a45bdcb14c3e1eefb0dcf91668d88ee"),
    ("Gold Blackjack 1", "2955688bdb4f23686e3ce61b905aafeb"),
    ("Gold Blackjack 3", "02478a653aa641470951d0a9cae59699"),
    ("Blackjack 1", "753fd4063959abf96f927fe171632d47"),
    ("VIP No Commission Speed Cricket Baccarat", "5619183cf03c3b03ebd01bbf42b37de4"),
    ("VIP Fortune Baccarat", "044a24737767690ed7a0be43ed9dd137"),
    ("Speed Fortune Baccarat", "a4f20ded65fffacd9001782619a90cce"),
    ("Fortune Baccarat", "04f266b7a2e9e68865d52fb7f2ac5e8a"),
    ("Speed Cricket Baccarat", "928af567b0839c6496bbfbb5709c5014"),
    ("Platinum Blackjack 1", "257b58d4471d0ea234380c10b145915f"),
    ("Gold Blackjack 4", "a75308c716157fde9e4faf84bcf80f1f"),
    ("Video Blackjack", "c9e306299b99ad529789673a6b4a8b88"),
    ("Unlimited Blackjack", "18cf7864fee424c7471bb7996aa4d37a"),
    ("Ultimate Sic Bo", "5cd59a9381764a84f5792d237469903a"),
    ("Sic Bo", "101e3c281b35485001bec47561a0a03e"),
    ("Russian Poker", "31ee0411b49acf83932fa0519e676997"),
    ("Italian Roulette", "2e31c310ad2491d3c6021f6063dc9b74"),
    ("EZ Dealer Roulette Thai", "a987ab0cda923c2f8e6fbc5292d7a062"),
    ("Portomaso Roulette 2", "91a2daa3d4b8065ffb75818568907ff8"),
    ("Casino Marina Baccarat 1", "2227c0b7445885e9f6a852eaf2fe74b6"),
    ("Casino Marina Andar Bahar", "5058f0aa42547208b1307fcbf21dcf9a"),
    ("One Day Teen Patti", "01556a46c5163d5570739dd7cddfcf68"),
    ("No Commission Baccarat", "958d30b26401872450a74a2d710adef6"),
    ("Lucky 7", "c88c40ec4fc544518d938315e2d1b2a3"),
    ("Dragon Tiger", "efdb52994fbfe97efcbd878dbd697ebb"),
    ("Cricket War", "93e289d1b18a9f82fb5d790f3c8e6735"),
    ("Casino Hold'em", "045e21f65e0e96eb502a4856ca9ababb"),
    ("Salon Prive Blackjack", "9f944a6cb336df7664f81e3ff6aba50e"),
    ("Gold Blackjack 5", "b307868469ec2b2e612045335086ff33"),
    ("Bet on Teen Patti", "e1b5650cd867be7719c15e7596aa7217"),
    ("Bet On Numbers HD", "531bac2b5726f8d9249eb5b0d432cb97"),
    ("Super 6 Baccarat", "f1fa68fce40959ce6ad5f367739f9e27"),
    ("Knockout Baccarat", "b12517092523e6d4b0c991a181c7d813"),
    ("Baccarat", "add11b218177a9898c594233148ca740"),
    ("Auto Roulette", "9b8fae325d15021b79f4fc650d5b8df5"),
    ("Andar Bahar", "435b892a73bf466e0ad584d480e12143"),
    ("32 Cards", "69e690d4f810d033fb4bb8ac7f3cc12f"),
    ("Ezugi Lobby", "d0e052b031dfcdb08d1803f4bcc618ef"),
]

JILI_GAMES: list[tuple[str, str]] = [
    ("Royal Fishing", "e794bf5717aca371152df192341fe68b"),
    ("Bombing Fishing", "e333695bcff28acdbecc641ae6ee2b23"),
    ("Dinosaur Tycoon", "eef3e28f0e3e7b72cbca61e7924d00f1"),
    ("Jackpot Fishing", "3cf4a85cb6dcf4d8836c982c359cd72d"),
    ("Dragon Fortune", "1200b82493e4788d038849bca884d773"),
    ("Mega Fishing", "caacafe3f64a6279e10a378ede09ff38"),
    ("Boom Legend", "f02ede19c5953fce22c6098d860dadf4"),
    ("Happy Fishing", "71c68a4ddb63bdc8488114a08e603f1c"),
    ("All-star Fishing", "9ec2a18752f83e45ccedde8dfeb0f6a7"),
    ("Dinosaur Tycoon II", "bbae6016f79f3df74e453eda164c08a4"),
    ("Ocean King Jackpot", "564c48d53fcddd2bcf0bf3602d86c958"),
    ("Chin Shi Huang", "24da72b49b0dd0e5cbef9579d09d8981"),
    ("God Of Martial", "21ef8a7ddd39836979170a2e7584e333"),
    ("Hot Chilli", "c845960c81d27d7880a636424e53964d"),
    ("Fortune Tree", "6a7e156ceec5c581cd6b9251854fe504"),
    ("War Of Dragons", "4b1d7ffaf9f66e6152ea93a6d0e4215b"),
    ("Gem Party", "756cf3c73a323b4bfec8d14864e3fada"),
    ("Lucky Ball", "893669898cd25d9da589a384f1d004df"),
    ("Hyper Burst", "a47b17970036b37c1347484cf6956920"),
    ("Shanghai Beauty", "795d0cae623cbf34d7f1aa93bbcded28"),
    ("Fa Fa Fa", "54c41adcf43fdb6d385e38bc09cd77ca"),
    ("Candy Baby", "2cc3b68cbcfacac2f7ef2fe19abc3c22"),
    ("Hawaii Beauty", "6409b758471b6df30c6b137b49f4d92e"),
    ("SevenSevenSeven", "61d46add6841aad4758288d68015eca6"),
    ("Bubble Beauty", "a78d2ed972aab8ba06181cc43c54a425"),
    ("FortunePig", "8488c76ee2afb8077fbd7eec62721215"),
    ("Crazy777", "8c62471fd4e28c084a61811a3958f7a1"),
    ("Bao boon chin", "8c4ebb3dc5dcf7b7fe6a26d5aadd2c3d"),
    ("Night City", "78e29705f7c6084114f46a0aeeea1372"),
    ("Fengshen", "09699fd0de13edbb6c4a194d7494640b"),
    ("Crazy FaFaFa", "a57a8d5176b54d4c825bd1eee8ab34df"),
    ("XiYangYang", "5a962d0e31e0d4c0798db5f331327e4f"),
    ("DiamondParty", "48d598e922e8c60643218ccda302af08"),
    ("Golden Bank", "c3f86b78938eab1b7f34159d98796e88"),
    ("Dragon Treasure", "c6955c14f6c28a6c2a0c28274fec7520"),
    ("Charge Buffalo", "984615c9385c42b3dad0db4a9ef89070"),
    ("Lucky Goldbricks", "d84ef530121953240116e3b2e93f6af4"),
    ("Super Ace", "bdfb23c974a2517198c5443adeea77a8"),
    ("Money Coming", "db249defce63610fccabfa829a405232"),
    ("Golden Queen", "8de99455c2f23f6827666fd798eb80ef"),
    ("Jungle King", "4db0ec24ff55a685573c888efed47d7f"),
    ("Monkey Party", "fd369a4a7486ff303beea267ec5c8eff"),
    ("Boxing King", "981f5f9675002fbeaaf24c4128b938d7"),
    ("Secret Treasure", "1d1f267e3a078ade8e5ccd56582ac94f"),
    ("Pharaoh Treasure", "c7a69ab382bd1ff0e6eb65b90a793bdd"),
    ("Lucky Coming", "ba858ec8e3b5e2b4da0d16b3a2330ca7"),
    ("Super Rich", "b92f491a63ac84b106b056e9d46d35c5"),
    ("RomaX", "e5ff8e72418fcc608d72ea21cc65fb70"),
    ("Golden Empire", "490096198e28f770a3f85adb6ee49e0f"),
    ("Fortune Gems", "a990de177577a2e6a889aaac5f57b429"),
    ("Crazy Hunter", "69082f28fcd46cbfd10ce7a0051f24b6"),
    ("Party Night", "d505541d522aa5ca01fc5e97cfcf2116"),
    ("Magic Lamp", "582a58791928760c28ec4cef3392a49f"),
    ("Agent Ace", "8a4b4929e796fda657a2d38264346509"),
    ("TWIN WINS", "c74b3cbda5d16f77523e41c25104e602"),
    ("Ali Baba", "cc686634b4f953754b306317799f1f39"),
    ("Mega Ace", "eba92b1d3abd5f0d37dfbe112abdf0e2"),
    ("Medusa", "2c17b7c4e2ce5b8bebf4bd10e3e958d7"),
    ("Book of Gold", "6b283c434fd44250d83b7c2420f164f9"),
    ("Thor X", "7e6aa773fa802aaa9cb1f2fac464736e"),
    ("Happy Taxi", "1ed896aae4bdc78c984021307b1dd177"),
    ("Gold Rush", "2a5d731e0fd60f52873a24ece11f2c0b"),
    ("Mayan Empire", "5c2383ef253f9c36dacec4b463d61622"),
    ("Crazy Pusher", "00d92d5cec10cf85623938222a6c2bb6"),
    ("Bone Fortune", "aab3048abc6a88e0759679fbe26e6a8d"),
    ("JILI CAISHEN", "11e330c2b23f106815f3b726d04e4316"),
    ("Bonus Hunter", "39775cdc4170e56c5f768bdee8b4fa00"),
    ("World Cup", "28374b7ad7c91838a46404f1df046e5a"),
    ("Samba", "6d35789b2f419c1db3926350d57c58d8"),
    ("Neko Fortune", "9a391758f755cb30ff973e08b2df6089"),
    ("Wild Racer", "2f0c5f96cda3c6e16b3929dd6103df8e"),
    ("Pirate Queen", "70999d5bcf2a1d1f1fb8c82e357317f4"),
    ("Golden Joker", "f301fe0b22d1540b1f215d282b20c642"),
    ("Wild Ace", "9a3b65e2ae5343df349356d548f3fc4b"),
    ("Master Tiger", "d2b48fe98ac2956eeefd2bc4f7e0335a"),
    ("Fortune Gems 2", "664fba4da609ee82b78820b1f570f4ad"),
    ("Sweet Land", "91250a55f75a3c67ed134b99bf587225"),
    ("Cricket King 18", "dcf220f4e3ecca0278911a55e6f11c77"),
    ("Elf Bingo", "5cec2b309a8845b38f8e9b4e6d649ea2"),
    ("Cricket Sah 75", "6720a0ce1d06648ff390fbea832798a9"),
    ("Golden Temple", "976c5497256c020ac012005f6bb166ad"),
    ("Devil Fire", "1b4c5865131b4967513c1ee90cba4472"),
    ("Bangla Beauty", "6b60d159f0939a45f7b4c88a9b57499a"),
    ("Aztec Priestess", "6acff19b2d911a8c695ba24371964807"),
    ("Fortune Monkey", "add95fc40f1ef0d56f5716ce45a56946"),
    ("Dabanggg", "5404a45b06826911c3537fdf935c281f"),
    ("Sin City", "830cac2f5da6cc1fb91cfae04b85b1e2"),
    ("King Arthur", "fafab1a17a237d0fc0e50c20d2c2bf4c"),
    ("Charge Buffalo Ascent", "28bc4a33c985ddce6acd92422626b76f"),
    ("Witches Night", "82c5c404cf4c0790deb42a2b5653533c"),
    ("Big Small", "25822eb4d6459cc8b39c4f7b69b1bf2c"),
    ("Number King", "36d20c24669dca7630715f2e0a7c18be"),
    ("Journey West M", "0d0a5a1731a6a05ffeb0e0f9d1948f80"),
    ("Poker King", "a9b13010273fcb0284c9ef436c5fe2ff"),
    ("Dragon & Tiger", "e7ac92d2fdd2aedca92a3521b4416f47"),
    ("iRich Bingo", "a53e46bf1e31f7a960ae314dc188e8b3"),
    ("7up7down", "3aca3084a5c1a8c77c52d6147ee3d2ab"),
    ("Baccarat", "b9c7c5f589cdaa63c4495e69eaa6dbbf"),
    ("Fortune Bingo", "2fd70535a3c838a438b4b8003ecce49d"),
    ("Sic Bo", "de0dc8a7fd369bd39a2d5747be87825c"),
    ("Super Bingo", "c934e67c2a84f52ef4fb598b56f3e7ba"),
    ("Bingo Carnaval", "d419ec9ab6a23590770fd77b036aed16"),
    ("Win Drop", "8211bc6e55e84d266bef9a6960940183"),
    ("Lucky Bingo", "c9f2470e285f3580cd761ba2e1f067e1"),
    ("Jackpot Bingo", "780d43c0a98bc8f6a0705976605608c3"),
    ("Color Game", "2ac4917fbc8b2034307b0c3cdd90d416"),
    ("Go Goal BIngo", "4e5ddaa644badc5f68974a65bf7af02a"),
    ("Calaca Bingo", "b2f05dae5370035a2675025953d1d115"),
    ("PAPPU", "e5091890bbb65a5f9ceb657351fa73c1"),
    ("West Hunter Bingo", "8d2c1506dc4ae4c47d23f9359d71c360"),
    ("Bingo Adventure", "2303867628a9a62272da7576665bbc65"),
    ("Golden Land", "05fc951a633d4c6b4bbe8c429cd63658"),
    ("Candyland Bingo", "711acbdf297ce40a09dd0e9023b63f50"),
    ("Color Prediction", "4a64504353c2304a3061bfd31cd9a62e"),
    ("Magic Lamp Bingo", "848ac1703885d5a86b54fbbf094b3b63"),
    ("Pearls of Bingo", "0995142f4685f66dfdd1a54fffa66ffa"),
    ("European Roulette", "d4fc911a31b3a61edd83bdd95e36f3bf"),
    ("Go Rush", "edef29b5eda8e2eaf721d7315491c51d"),
    ("Mines", "72ce7e04ce95ee94eef172c0dfd6dc17"),
    ("Tower", "8e939551b9e785001fcb5b0a32f88aba"),
    ("HILO", "bd8a2bb2dd63503b93cf6ac9492786ce"),
    ("Limbo", "eabf08253165b6bb2646e403de625d1a"),
    ("Wheel", "6e19e03c50f035ddd9ffd804c30f8c80"),
    ("Mines Gold", "4bceeb28b1a88c87d1ef518d7af2bba9"),
    ("Keno", "a54e3f5e231085c7d8ba99e8ed2261fc"),
    ("Plinko", "e3b71c6844eb8c30f5ef210ad92725a6"),
    ("Crash Bonus", "a7f3e5f210523a989a7c6b32f2f1ad42"),
    ("TeenPatti", "f743cb55c2c4b737727ef144413937f4"),
    ("AK47", "488c377662cad37a551bde18e2fbe785"),
    ("Andar Bahar", "6f48b3aa0b64c79a2dc320ea021148b5"),
    ("Rummy", "ae632f32c3a1e6803f9a6fbec16be28e"),
    ("Callbreak", "9092b5a56e001c60850c4c1184c53e07"),
    ("TeenPatti Joker", "1a4eaca67612e65fdcae43f4c8a667a4"),
    ("Callbreak Quick", "aa9a9916d6e48ba50afa3c2246b6dacb"),
    ("TeenPatti 20-20", "1afa7db588d05de7b9abca4664542765"),
    ("Ludo Quick", "bb1f14d788d37b06dc8f6701ed57ed0d"),
    ("Tongits Go", "26fbfab92a3837b7dbf767e783b173af"),
    ("Pusoy Go", "f2879a3f20f305eadad13448e11c052e"),
    ("Blackjack", "3b502aee6c9e1ef0f698332ee1b76634"),
    ("Blackjack Lucky Ladies", "d0d1c20062e28493e1750f27a1730c48"),
    ("MINI FLUSH", "07afefc388ab6af8cf26f85286f83fae"),
    ("Pool Rummy", "43e7df819bf57722a8917bb328640b30"),
    ("Caribbean Stud Poker", "04c9784b0b1b162b2c86f9ce353da8b7"),
    ("Fortune Gems 3", "63927e939636f45e9d6d0b3717b3b1c1"),
    ("Super Ace Deluxe", "80aad2a10ae6a95068b50160d6c78897"),
    ("3 Coin Treasures", "69c1b4586b5060eefcb45bb479f03437"),
    ("3 Lucky Piggy", "e09d4c9612ea540bc0afabf76e4f9148"),
    ("Poseidon", "50a1bcbc2ef4a5f761e0e4d338a41699"),
    ("3 Pot Dragons", "921dce2d616e5d0577135bb2d9214946"),
    ("Money Pot", "a5acbbb7ae534d303f67cb447dc8723d"),
    ("Nightfall Hunting", "ced5e3de03293fc6fb111298a504cfeb"),
    ("Shōgun", "68724804a3cd30c749e460256b462f00"),
    ("Ultimate Texas Hold'em", "82fa04ccbbf20291128408c014092bce"),
    ("Devil Fire 2", "0426ba674c9dd29de6fa023afcf0640d"),
    ("Legacy of Egypt", "1310248a5eab24b4bf113a6e0ee7962a"),
    ("Lucky Jaguar", "731e642b1fee94725e7313f3dfba8f45"),
    ("Jackpot Joker", "7ed860eef313538545ff7aa2b9290cf9"),
    ("Fortune King Jackpot", "f2b04833d555ef9989748f9ecabd5249"),
    ("Arena Fighter", "71468f38b1fa17379231d50635990c31"),
    ("Trial of Phoenix", "d11ea63b63ec615ae6df589f0b0d53e1"),
    ("Zeus", "4e7c9f4fbe9b5137f21ebd485a9cfa5c"),
    ("Potion Wizard", "fba154365cdf8fad07565cf93bae3521"),
    ("Crazy Hunter 2", "68880d1fcbd274f6b2bf7168276af51d"),
    ("The Pig House", "824736d3e6abff8a0b7e79d784c7b113"),
    ("Money Coming Expand Bets", "3a557646c3abb12201c0b8810a8c0966"),
    ("Party Star", "bfde2986a4eb3a5a559ac8a8c64df461"),
    ("Egypts Glow", "ddac017cb273a590b7aa0e1ad6a52bef"),
    ("Lucky Doggy", "4bf1d6a75d91c725f89aa5985544a087"),
    ("Golden Bank 2", "3a72a27c8851be5a396f51a19654c7c3"),
    ("Fruity Wheel", "921cf987632d65b5e41ab5dffe16d95a"),
    ("Safari Mystery", "56dad0ca19e96dc6ee1038d374712767"),
    ("Treasure Quest", "6bb74b0a57a66850b79ab5c93864cac3"),
    ("Coin Tree", "ca72a7ad1ca4fa2cdc9a1c49c8bb3332"),
    ("3 Coin Wild Horse", "25bff08b69ccd31c238a627b53afff36"),
    ("Jogo do Bicho", "29149ed003ec05873fc164fc139b5606"),
    ("Speed Baccarat", "9e969a7e77e8f61dbe94575e6c96272f"),
    ("Fish Prawn Crab", "a231addcfef742c0c55049a0cde6e674"),
    ("Super Ace Scratch", "0ec0aeb7aad8903bb6ee6b9b9460926a"),
    ("Jhandi Munda", "3a2f7e03e9e86c925ab8c8612f2ea259"),
    ("Boxing Extravaganza", "2469a7a4bdf296f841f59d0a42bba1a8"),
    ("Thai Hilo", "db89731adb091f081381d77eb5a06162"),
    ("Super E-Sabong", "706342709fa0e5f40068e4e6d81f7358"),
    ("Cricket Roulette", "b761552c4191ee73c2c323e34883b57a"),
    ("Fortune Roulette", "858afeefec569d30eb8a041b335e7507"),
    ("Go For Champion", "45a2a090dd3f8c5e51a20e5f7c24830b"),
    ("Fortune Gems Scratch", "d528913b832aba97654b6393b3a915b4"),
    ("Crash Goal", "a8a5f7f458c8507c311ed7e396fdddd8"),
    ("Crash Cricket", "2b3e6d31620ceabbe44d3edb6fd10af3"),
    ("Keno Bonus Number", "7a24f38a556e6c0682d9ad4f22f60452"),
    ("Video Poker", "28d459e6b8bba9a375e65e1f25e8d316"),
    ("Super Ace Joker", "29c66f73e3916b8eb18c2bf78886927d"),
    ("Pirate Queen 2", "4702eb871271aa62ef3f3d78f5d968c1"),
    ("Coin Infinity Surge Reel", "a1ea10a6b30f260b6d6ff17028d38913"),
    ("Sweet Magic", "ae88afcb58415b7802e2c02c40816f17"),
    ("3 Charge Buffalo", "3ea8ed5f8ba2239e6cd49366afb743f8"),
    ("3 Rich pigies", "472f684f667e272e0ccc7ac1529170ca"),
    ("Roma X Deluxe", "b4fe8cea772a7643551a12de806472e8"),
    ("3 LUCKY LION", "7af6be9d29bb593fa0f6516b14b02103"),
    ("Bikini Lady", "702565a827764d10e470a0f76398a978"),
    ("Fortune Coins", "d6d14943efe13dd3bcf1428d0f702024"),
    ("Cricket War", "a4224fdff8b66bd55ab891e2fd879ac1"),
    ("Keno Super Chance", "8f0ea9429cab15f2a48d9f4972d30b52"),
    ("Crash Touchdown", "014c49675e1c22c76352b8047ae6d8eb"),
    ("Keno Extra Bet", "191d02a6e852cd18ce1dd4d175e96cd6"),
    ("3 Coin Treasures 2", "7b4308e95fa25021bae874f9e128c8c3"),
    ("Penalty Kick", "446de3502193f08cdf0c17bf0791eb41"),
]

# ---------------------------------------------------------------------------
# Provider registry — maps provider_code -> (display_name, image_folder_slug)
# image_folder_slug: key passed to get_image_folder_candidates; None = no images
# ---------------------------------------------------------------------------

PROVIDERS: dict[str, tuple[str, str | None]] = {
    "ezugi":             ("Ezugi",              "ezugi"),
    "jili":              ("JILI",               "jili"),
    "evolution_live":    ("Evolution Live",     "evolution_live"),
    "evoplay_asia":      ("Evoplay Asia",       None),
    "pragmatic_live":    ("Pragmatic Live",     "pragmatic_live"),
    "saba_sports":       ("SABA Sports",        None),
    "sexy_gaming":       ("Sexy Gaming",        "sexy_gaming"),
    "smartsoft_gaming":  ("SmartSoft Gaming",   None),
    "spribe":            ("Spribe",             "spribe"),
    "lucksportsgaming":  ("LuckSportsGaming",   None),
}

# JILI API config
JILI_API_ENDPOINT = "https://allapi.online/launch_game1_js"
JILI_API_SECRET = "4d45bba519ac2d39d1618f57120b84b7"
JILI_API_TOKEN = "184de030-912d-4c26-81fc-6c5cd3c05add"

# ---------------------------------------------------------------------------
# GameRow: (provider_code, game_name, game_uid)
# ---------------------------------------------------------------------------

GameRow = tuple[str, str, str]


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _normalize_slug(code: str) -> str:
    s = (code or "").strip().lower()
    return "".join(c if c.isalnum() or c in " -_" else "" for c in s).replace(" ", "_").replace("-", "_").strip("_") or "unknown"


def load_embedded() -> list[GameRow]:
    rows: list[GameRow] = []
    for name, uid in EZUGI_GAMES:
        if name and uid:
            rows.append(("ezugi", name.strip()[:255], uid.strip()[:255]))
    for name, uid in JILI_GAMES:
        if name and uid:
            rows.append(("jili", name.strip()[:255], uid.strip()[:255]))
    return rows


def load_spribe_txt(docs_games: Path) -> list[GameRow]:
    """spribe.txt: 4-col tab-separated — index, provider, name, uid."""
    rows: list[GameRow] = []
    path = docs_games / "spribe.txt"
    if not path.exists():
        return rows
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return rows
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.replace(",", "\t").split("\t")]
        if len(parts) >= 4 and parts[3]:
            rows.append(("spribe", parts[2][:255] or parts[3][:255], parts[3][:255]))
    return rows


def load_lucksports_txt(docs_games: Path) -> list[GameRow]:
    """lucksportsgaming.txt: 2-col — provider, uid (game name = uid)."""
    rows: list[GameRow] = []
    path = docs_games / "lucksportsgaming.txt"
    if not path.exists():
        return rows
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return rows
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.replace(",", "\t").split("\t")]
        if len(parts) >= 2 and parts[1]:
            rows.append(("lucksportsgaming", parts[1][:255], parts[1][:255]))
        elif len(parts) == 1 and parts[0]:
            rows.append(("lucksportsgaming", parts[0][:255], parts[0][:255]))
    return rows


def load_xlsx_with_header(path: Path, provider_code: str) -> list[GameRow]:
    """
    XLSX with header row. Expected columns (case-insensitive, spaces/underscores stripped):
      game_name / gamename / name  -> game display name
      game_id / gameid / uid       -> unique game id
    Falls back to column indices if header detection fails.
    """
    rows: list[GameRow] = []
    try:
        import openpyxl
    except ImportError:
        return rows
    if not path.exists():
        return rows
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return rows
    try:
        sheet = wb.active
        if not sheet:
            return rows
        header = [str(c.value or "").strip().lower().replace(" ", "").replace("_", "") for c in sheet[1]]

        def col(aliases: list[str]) -> int | None:
            for alias in aliases:
                for i, h in enumerate(header):
                    if alias == h or alias in h:
                        return i
            return None

        name_col = col(["gamename", "name", "game"])
        uid_col = col(["gameid", "gameuid", "uid", "gameid"])
        # If header detection fails, try positional defaults (id, game_api, game_name, game_id)
        if name_col is None:
            name_col = 2
        if uid_col is None:
            uid_col = 3

        for row in sheet.iter_rows(min_row=2):
            values = [str(c.value or "").strip() for c in row]
            if not values or uid_col >= len(values):
                continue
            uid = values[uid_col][:255] if values[uid_col] else ""
            name = values[name_col][:255] if name_col < len(values) else ""
            if not uid or not name:
                continue
            rows.append((provider_code, name, uid))
    finally:
        wb.close()
    return rows


def load_xlsx_no_header(path: Path, provider_code: str) -> list[GameRow]:
    """
    XLSX without header. Col 0 = game name, col 1 = uid.
    Used for SABA Sports and Sexy Gaming.
    """
    rows: list[GameRow] = []
    try:
        import openpyxl
    except ImportError:
        return rows
    if not path.exists():
        return rows
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return rows
    try:
        sheet = wb.active
        if not sheet:
            return rows
        for row in sheet.iter_rows(min_row=1):
            values = [str(c.value or "").strip() for c in row]
            if len(values) < 2 or not values[1]:
                continue
            name = values[0][:255]
            uid = values[1][:255]
            if not name:
                continue
            rows.append((provider_code, name, uid))
    finally:
        wb.close()
    return rows


def _find_file_case_insensitive(folder: Path, filename: str) -> Path:
    """
    Return the path for filename inside folder, even if the actual file on disk
    uses different casing.  Falls back to the original (non-existing) path so
    callers can still check .exists() normally.
    """
    target = folder / filename
    if target.exists():
        return target
    # Scan directory entries with a case-insensitive comparison
    lower = filename.lower()
    try:
        for entry in folder.iterdir():
            if entry.name.lower() == lower:
                return entry
    except Exception:
        pass
    return target  # original path; callers will see .exists() == False


def load_all_xlsx(docs_games: Path) -> list[tuple[str, str, list[GameRow], bool]]:
    """
    Returns a list of (filename, provider_code, rows, found) tuples so callers
    can emit per-source logging.  found=True when the file was located on disk.
    """
    results: list[tuple[str, str, list[GameRow], bool]] = []
    # XLSX files with header row (col names in row 1)
    header_files = {
        "Evolution live.xlsx":    "evolution_live",
        "evoplay asia.xlsx":      "evoplay_asia",
        "Pragmatic live.xlsx":    "pragmatic_live",
        "SmartSoft Gaming.xlsx":  "smartsoft_gaming",
    }
    # XLSX files without header (data starts at row 1; col0=name, col1=uid)
    no_header_files = {
        "SABA Sports.xlsx":  "saba_sports",
        "Sexy Gaming.xlsx":  "sexy_gaming",
    }
    for filename, code in header_files.items():
        path = _find_file_case_insensitive(docs_games, filename)
        rows = load_xlsx_with_header(path, code)
        results.append((filename, code, rows, path.exists()))
    for filename, code in no_header_files.items():
        path = _find_file_case_insensitive(docs_games, filename)
        rows = load_xlsx_no_header(path, code)
        results.append((filename, code, rows, path.exists()))
    return results


# ---------------------------------------------------------------------------
# Main command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = (
        "Super Game Seeder: seeds all 10 providers, per-game categories "
        "(inferred from game name/provider), and games with images. "
        "Sources: embedded Ezugi/JILI lists + docs/games XLSX & TXT files. "
        "--full-reset: wipe all and re-seed. --fresh: wipe only seeded providers/games."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print what would be created; no DB writes.",
        )
        parser.add_argument(
            "--fresh",
            action="store_true",
            help="Delete only providers/games that would be seeded, then re-seed.",
        )
        parser.add_argument(
            "--full-reset",
            action="store_true",
            help="Delete ALL Game, GameCategory, GameProvider, then re-seed.",
        )
        parser.add_argument(
            "--providers",
            type=str,
            default="",
            help="Comma-separated provider codes to limit seeding (e.g. ezugi,jili,spribe).",
        )
        parser.add_argument(
            "--images-only",
            action="store_true",
            help="Only fill missing game images; do not create new games.",
        )

    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        dry_run    = options.get("dry_run", False)
        fresh      = options.get("fresh", False)
        full_reset = options.get("full_reset", False)
        images_only = options.get("images_only", False)
        providers_filter = [
            p.strip().lower()
            for p in (options.get("providers") or "").split(",")
            if p.strip()
        ]

        docs_games = DOCS_GAMES
        if not docs_games.exists() and not dry_run:
            self.stdout.write(
                self.style.WARNING(
                    f"docs/games path not found: {docs_games}. "
                    "Place XLSX/TXT and image folders there, or set DOCS_GAMES_PATH (env or settings)."
                )
            )

        # ── Load all game rows ──────────────────────────────────────────
        self.stdout.write("Loading game data sources:")
        all_rows: list[GameRow] = []

        embedded = load_embedded()
        all_rows.extend(embedded)
        ezugi_count = sum(1 for r in embedded if r[0] == "ezugi")
        jili_count  = sum(1 for r in embedded if r[0] == "jili")
        self.stdout.write(f"  [embedded]         ezugi={ezugi_count} games, jili={jili_count} games")

        spribe_rows = load_spribe_txt(docs_games)
        all_rows.extend(spribe_rows)
        spribe_path = docs_games / "spribe.txt"
        status = "OK" if spribe_path.exists() else "MISSING"
        self.stdout.write(f"  [spribe.txt]       {len(spribe_rows)} games  [{status}]")

        lucks_rows = load_lucksports_txt(docs_games)
        all_rows.extend(lucks_rows)
        lucks_path = docs_games / "lucksportsgaming.txt"
        status = "OK" if lucks_path.exists() else "MISSING"
        self.stdout.write(f"  [lucksportsgaming.txt] {len(lucks_rows)} games  [{status}]")

        for filename, code, xlsx_rows, found in load_all_xlsx(docs_games):
            all_rows.extend(xlsx_rows)
            status = "OK" if found else "MISSING – place file in docs/games/"
            self.stdout.write(f"  [{filename:<30}] {len(xlsx_rows):>3} games  [{status}]")

        if providers_filter:
            all_rows = [r for r in all_rows if r[0].lower() in providers_filter]

        if not all_rows:
            self.stdout.write(self.style.WARNING("No game rows found (check docs/games and filters)."))
            return

        self.stdout.write(
            self.style.SUCCESS(
                f"Loaded {len(all_rows)} game rows across "
                f"{len({r[0] for r in all_rows})} providers: "
                f"{sorted({r[0] for r in all_rows})}"
            )
        )

        if dry_run:
            unique = {(r[0], r[2]) for r in all_rows}
            self.stdout.write(self.style.SUCCESS(
                f"Dry run: would seed {len(unique)} unique games for providers: "
                f"{sorted({r[0] for r in all_rows})}."
            ))
            return

        # ── Full reset ──────────────────────────────────────────────────
        if full_reset and not images_only:
            gc, _ = Game.objects.all().delete()
            cc, _ = GameCategory.objects.all().delete()
            pc, _ = GameProvider.objects.all().delete()
            self.stdout.write(self.style.WARNING(
                f"Full reset: deleted {gc} games, {cc} categories, {pc} providers."
            ))

        # ── Fresh: delete only affected providers/games ─────────────────
        elif fresh and not images_only:
            affected_codes = list({r[0] for r in all_rows})
            for code in affected_codes:
                prov = GameProvider.objects.filter(code=code).first()
                if prov:
                    d, _ = Game.objects.filter(provider=prov).delete()
                    prov.delete()
                    self.stdout.write(self.style.WARNING(f"Fresh: deleted provider '{code}' and {d} games."))

        # ── Step 1: Category cache (populated dynamically per game) ─────
        cat_cache: dict[str, GameCategory] = {}

        def get_cat(cat_name: str) -> GameCategory:
            cat_name = (cat_name or "Other").strip()[:255] or "Other"
            if cat_name in cat_cache:
                return cat_cache[cat_name]
            obj, created = GameCategory.objects.get_or_create(
                name=cat_name,
                defaults={"is_active": True},
            )
            cat_cache[cat_name] = obj
            if created:
                self.stdout.write(f"  Category created: {cat_name}")
            return obj

        # ── Step 2: Seed providers ──────────────────────────────────────
        prov_cache: dict[str, GameProvider] = {}
        provider_codes_in_data = {r[0] for r in all_rows}
        for code in provider_codes_in_data:
            display_name, _ = PROVIDERS.get(code, (code.replace("_", " ").title(), None))
            defaults: dict = {"name": display_name, "is_active": True}
            if code == "jili":
                defaults["api_endpoint"] = JILI_API_ENDPOINT
                defaults["api_secret"]   = JILI_API_SECRET
                defaults["api_token"]    = JILI_API_TOKEN
            prov_obj, created = GameProvider.objects.get_or_create(code=code, defaults=defaults)
            prov_cache[code] = prov_obj
            if created:
                self.stdout.write(f"  Provider created: {display_name} ({code})")

        # ── Step 3: Seed games ──────────────────────────────────────────
        zero = Decimal("0")
        created_games = 0
        skipped_games = 0
        images_set = 0

        def get_image(provider_code: str, game_name: str, game_uid: str) -> Path | None:
            _, img_slug = PROVIDERS.get(provider_code, (None, None))
            if img_slug is None:
                return None
            candidates = list(get_image_folder_candidates(docs_games, img_slug))
            # For JILI, also check docs/games/jiliwebp directly
            if provider_code == "jili":
                jili_direct = docs_games / "jiliwebp"
                if jili_direct not in candidates:
                    candidates.append(jili_direct)
            return find_image_for_game_in_folders(candidates, game_name, game_uid)

        for provider_code, game_name, game_uid in all_rows:
            provider = prov_cache.get(provider_code)
            if not provider:
                # Fallback: look up in DB (handles edge cases after --fresh)
                provider = GameProvider.objects.filter(code=provider_code).first()
                if not provider:
                    self.stdout.write(self.style.WARNING(f"Provider '{provider_code}' not found, skipping {game_name!r}."))
                    continue
                prov_cache[provider_code] = provider

            cat_name = infer_subcategory(game_name, provider_code)
            cat      = get_cat(cat_name)

            if images_only:
                game = Game.objects.filter(provider=provider, game_uid=game_uid).first()
                if game and not game.image:
                    img_path = get_image(provider_code, game_name, game_uid)
                    if img_path:
                        try:
                            with open(img_path, "rb") as f:
                                game.image.save(img_path.name, ContentFile(f.read()), save=True)
                            images_set += 1
                        except Exception as e:
                            self.stdout.write(self.style.WARNING(f"Image save failed for {game_name!r}: {e}"))
                continue

            game, was_created = Game.objects.get_or_create(
                provider=provider,
                game_uid=game_uid,
                defaults={
                    "name":      game_name,
                    "category":  cat,
                    "is_active": True,
                    "min_bet":   zero,
                    "max_bet":   zero,
                },
            )

            if was_created:
                created_games += 1
            else:
                skipped_games += 1
                # Keep category in sync if it changed
                if game.category_id != cat.id:
                    game.category = cat
                    game.save(update_fields=["category"])

            if was_created or not game.image:
                img_path = get_image(provider_code, game_name, game_uid)
                if img_path:
                    try:
                        with open(img_path, "rb") as f:
                            game.image.save(img_path.name, ContentFile(f.read()), save=True)
                        images_set += 1
                    except Exception as e:
                        self.stdout.write(self.style.WARNING(f"Image save failed for {game_name!r}: {e}"))

        self.stdout.write(self.style.SUCCESS(
            f"\nDone! "
            f"Providers in data: {len(provider_codes_in_data)}. "
            f"Categories: {len(cat_cache)}. "
            f"Games created: {created_games}, skipped: {skipped_games}. "
            f"Images set: {images_set}."
        ))
